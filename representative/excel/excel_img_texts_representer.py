from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

import representative.report_representer as r_representer
import models.report_response as r_resp
from models.repo_documents import ImgTextsDocument, ImgTextsDocumentWithPath


class FromHTMLImgTextsRepresnter(r_representer.EXCELReportRepresenter):
    def __init__(self, img_texts_document_with_path: ImgTextsDocumentWithPath):
        self.datas: List[ImgTextsDocument] = img_texts_document_with_path.data
        self.path_to_save = img_texts_document_with_path.path

    def generate(self) -> r_resp.ReportResponse:
        wb = Workbook()
        ws: Worksheet = wb.active

        alignment = Alignment(horizontal='general',
                              vertical = 'top',
                              text_rotation = 0,
                              wrap_text = False,
                              shrink_to_fit = False,
                              indent = 0)

        indexes = ["A", "B", "C", "D", "E"]
        colls = ["File name", "Frame time", "Title", "Data"]
        for i in range(1, len(colls) + 1):
            ws[f'{indexes[i - 1]}{1}'] = colls[i - 1]

        ws.row_dimensions[1].height = 20

        for row_number in range(2,len(self.datas) + 2):
            ws.row_dimensions[row_number].height = 170
            doc = self.datas[row_number - 2]
            title_data = ' '.join(t['text'] for t in doc.header)
            doc_data = [doc.video_name, doc.frame_time.split('.')[0], title_data, doc.get_data_as_key_value_str]
            for coll_number in range(1, len(doc_data) + 1):
                cell = ws[f'{indexes[coll_number - 1]}{row_number}']
                cell.value = doc_data[coll_number - 1]
                cell.alignment = alignment

        for letter in indexes:
            ws.column_dimensions[letter].auto_size = True #.width = 30
            # ws.column_dimensions[letter].bestFit = True


        # for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
        #     for cell in rows:
        #         cell.fill = PatternFill(bgColor="00CCFFFF", fill_type="solid")

        wb.save(self.path_to_save)
        resp = "true"
        return r_resp.ReportResponse(resp)
